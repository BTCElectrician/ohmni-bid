#!/usr/bin/env python3
"""
Extract Pricing from Excel Master File

Reads the password-protected Excel file and extracts all pricing data
into pricing_database.json for use by the estimation system.

Usage:
    python scripts/extract_pricing.py

Requires:
    - .env file with EXCEL_PASSWORD
    - data/Unit price - Master 11-3-21- JG.xlsx
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    import pandas as pd
    from openpyxl import load_workbook
    from dotenv import load_dotenv
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip install pandas openpyxl python-dotenv")
    sys.exit(1)

# Load environment variables
load_dotenv(PROJECT_ROOT / ".env")

# Configuration
EXCEL_FILENAME = os.getenv("EXCEL_FILENAME", "Unit price - Master 11-3-21- JG.xlsx")
EXCEL_PASSWORD = os.getenv("EXCEL_PASSWORD", "")
EXCEL_PATH = PROJECT_ROOT / "data" / EXCEL_FILENAME
OUTPUT_PATH = PROJECT_ROOT / "flask_integration" / "data" / "pricing_database.json"


def extract_conduit_pricing(wb) -> list:
    """Extract conduit pricing from FEEDERS sheet."""
    conduit_items = []

    try:
        ws = wb["FEEDERS"]
    except KeyError:
        print("Warning: FEEDERS sheet not found")
        return conduit_items

    # Conduit type mappings based on row ranges
    conduit_types = [
        {"type": "EMT_SS", "typeName": "EMT Set Screw", "start_row": 10, "end_row": 19},
        {"type": "EMT_COMP", "typeName": "EMT Compression", "start_row": 22, "end_row": 31},
        {"type": "HW", "typeName": "Heavy Wall/Rigid", "start_row": 34, "end_row": 43},
        {"type": "IMC", "typeName": "IMC", "start_row": 46, "end_row": 55},
        {"type": "PVC", "typeName": "PVC Schedule 40", "start_row": 58, "end_row": 68},
        {"type": "PVC_GRC", "typeName": "PVC/GRC Assembly", "start_row": 71, "end_row": 80},
    ]

    sizes = ["1/2", "3/4", "1", "1-1/4", "1-1/2", "2", "2-1/2", "3", "3-1/2", "4", "5"]

    for ct in conduit_types:
        row = ct["start_row"]
        size_idx = 0
        while row <= ct["end_row"] and size_idx < len(sizes):
            try:
                # Column D = Material cost per 100ft, Column G = Labor hours per 100ft
                material_cost = ws.cell(row=row, column=4).value
                labor_hours = ws.cell(row=row, column=7).value

                if material_cost and labor_hours:
                    conduit_items.append({
                        "type": ct["type"],
                        "typeName": ct["typeName"],
                        "name": f"{sizes[size_idx]}\" {ct['type'].replace('_', ' ')} Unit",
                        "size": sizes[size_idx],
                        "materialCostPer100ft": float(material_cost) if material_cost else 0,
                        "laborHoursPer100ft": float(labor_hours) if labor_hours else 0,
                    })
            except Exception as e:
                print(f"Warning: Error reading conduit row {row}: {e}")

            row += 1
            size_idx += 1

    print(f"  Extracted {len(conduit_items)} conduit items")
    return conduit_items


def extract_wire_pricing(wb) -> list:
    """Extract wire pricing from FEEDERS sheet."""
    wire_items = []

    try:
        ws = wb["FEEDERS"]
    except KeyError:
        return wire_items

    # Wire sizes in order
    copper_sizes = ["#12", "#10", "#8", "#6", "#4", "#3", "#2", "#1",
                    "#1/0", "#2/0", "#3/0", "#4/0",
                    "250 MCM", "300 MCM", "350 MCM", "400 MCM", "500 MCM", "600 MCM", "750 MCM"]

    # Copper THHN: rows 83-101
    for i, size in enumerate(copper_sizes):
        row = 83 + i
        if row > 101:
            break
        try:
            market_price = ws.cell(row=row, column=3).value  # Column C
            markup = ws.cell(row=row, column=4).value  # Column D (as decimal)
            labor_hours = ws.cell(row=row, column=7).value  # Column G

            if market_price:
                # Determine markup based on size
                if size in ["#12", "#10", "#8"]:
                    markup_pct = 0.15
                elif size in ["#6", "#4"]:
                    markup_pct = 0.10
                elif size == "750 MCM":
                    markup_pct = 0.10
                else:
                    markup_pct = 0.05

                wire_items.append({
                    "material": "CU",
                    "type": "THHN",
                    "size": size,
                    "marketPricePer1000ft": float(market_price) if market_price else 0,
                    "markupPercent": markup_pct,
                    "laborHoursPer1000ft": float(labor_hours) if labor_hours else 0,
                })
        except Exception as e:
            print(f"Warning: Error reading wire row {row}: {e}")

    # Aluminum THHN: rows 108-120
    aluminum_sizes = ["#6", "#4", "#2", "#1/0", "#2/0", "#3/0", "#4/0",
                      "250 MCM", "300 MCM", "350 MCM", "400 MCM", "500 MCM", "600 MCM"]

    for i, size in enumerate(aluminum_sizes):
        row = 108 + i
        if row > 120:
            break
        try:
            market_price = ws.cell(row=row, column=3).value
            labor_hours = ws.cell(row=row, column=7).value

            if market_price:
                wire_items.append({
                    "material": "AL",
                    "type": "THHN",
                    "size": size,
                    "marketPricePer1000ft": float(market_price) if market_price else 0,
                    "markupPercent": 0.05,
                    "laborHoursPer1000ft": float(labor_hours) if labor_hours else 0,
                })
        except Exception as e:
            print(f"Warning: Error reading aluminum wire row {row}: {e}")

    print(f"  Extracted {len(wire_items)} wire items")
    return wire_items


def extract_line_items(wb) -> list:
    """Extract line items from basebid sheet."""
    line_items = []

    try:
        ws = wb["basebid"]
    except KeyError:
        print("Warning: basebid sheet not found")
        return line_items

    # Category mappings based on row ranges
    categories = [
        {"name": "TEMP_POWER", "start_row": 11, "end_row": 14},
        {"name": "ELECTRICAL_SERVICE", "start_row": 15, "end_row": 158},
        {"name": "MECHANICAL_CONNECTIONS", "start_row": 159, "end_row": 190},
        {"name": "INTERIOR_LIGHTING", "start_row": 191, "end_row": 231},
        {"name": "EXTERIOR_LIGHTING", "start_row": 232, "end_row": 247},
        {"name": "POWER_RECEPTACLES", "start_row": 248, "end_row": 285},
        {"name": "SITE_CONDUITS", "start_row": 286, "end_row": 294},
        {"name": "SECURITY", "start_row": 295, "end_row": 297},
        {"name": "FIRE_ALARM", "start_row": 299, "end_row": 311},
        {"name": "GENERAL_CONDITIONS", "start_row": 312, "end_row": 320},
    ]

    for cat in categories:
        for row in range(cat["start_row"], cat["end_row"] + 1):
            try:
                description = ws.cell(row=row, column=1).value  # Column A
                material_cost = ws.cell(row=row, column=4).value  # Column D
                unit_type = ws.cell(row=row, column=5).value  # Column E
                labor_hours = ws.cell(row=row, column=7).value  # Column G

                # Skip empty rows or header rows
                if not description or description == "MATERIAL":
                    continue
                if not material_cost and not labor_hours:
                    continue

                # Normalize unit type
                unit = str(unit_type).upper() if unit_type else "E"
                if unit not in ["E", "C", "M", "LOT"]:
                    unit = "E"

                line_items.append({
                    "category": cat["name"],
                    "description": str(description).strip(),
                    "materialUnitCost": float(material_cost) if material_cost else 0,
                    "laborHoursPerUnit": float(labor_hours) if labor_hours else 0,
                    "unitType": unit,
                })
            except Exception as e:
                # Skip problematic rows silently
                pass

    print(f"  Extracted {len(line_items)} line items")
    return line_items


def extract_parameters(wb) -> dict:
    """Extract global parameters from basebid sheet."""
    params = {
        "laborRate": 118.0,
        "materialTaxRate": 0.1025,
        "defaultOverheadProfitRate": 0
    }

    try:
        ws = wb["basebid"]

        # Labor rate: I323
        labor_rate = ws.cell(row=323, column=9).value
        if labor_rate:
            params["laborRate"] = float(labor_rate)

        # Material tax rate: E323
        tax_rate = ws.cell(row=323, column=5).value
        if tax_rate:
            params["materialTaxRate"] = float(tax_rate)

        # O&P rate: C329
        op_rate = ws.cell(row=329, column=3).value
        if op_rate:
            params["defaultOverheadProfitRate"] = float(op_rate)

    except Exception as e:
        print(f"Warning: Error extracting parameters: {e}")

    print(f"  Parameters: ${params['laborRate']}/hr, {params['materialTaxRate']*100}% tax")
    return params


def main():
    print("=" * 60)
    print("OHMNI BID - Pricing Extraction")
    print("=" * 60)

    # Check for Excel file
    if not EXCEL_PATH.exists():
        print(f"\nError: Excel file not found at {EXCEL_PATH}")
        print("Make sure the file is in the data/ folder")
        sys.exit(1)

    print(f"\nSource: {EXCEL_FILENAME}")
    print(f"Password: {'*' * len(EXCEL_PASSWORD) if EXCEL_PASSWORD else '(none)'}")

    # Load workbook
    print("\nLoading workbook...")
    try:
        if EXCEL_PASSWORD:
            # For password-protected files, we need to use msoffcrypto
            try:
                import msoffcrypto
                import io

                with open(EXCEL_PATH, "rb") as f:
                    decrypted = io.BytesIO()
                    office_file = msoffcrypto.OfficeFile(f)
                    office_file.load_key(password=EXCEL_PASSWORD)
                    office_file.decrypt(decrypted)
                    wb = load_workbook(decrypted, data_only=True)
            except ImportError:
                print("Note: msoffcrypto not installed, trying without password...")
                wb = load_workbook(EXCEL_PATH, data_only=True)
        else:
            wb = load_workbook(EXCEL_PATH, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        sys.exit(1)

    print(f"Sheets found: {wb.sheetnames}")

    # Extract data
    print("\nExtracting data...")

    parameters = extract_parameters(wb)
    conduit = extract_conduit_pricing(wb)
    wire = extract_wire_pricing(wb)
    line_items = extract_line_items(wb)

    # Build output
    output = {
        "version": "1.0",
        "source": EXCEL_FILENAME,
        "extracted_at": datetime.now().isoformat(),
        "parameters": parameters,
        "conduit": conduit,
        "wire": wire,
        "lineItems": line_items,
    }

    # Ensure output directory exists
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    # Write JSON
    print(f"\nWriting to {OUTPUT_PATH}...")
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)

    # Summary
    print("\n" + "=" * 60)
    print("EXTRACTION COMPLETE")
    print("=" * 60)
    print(f"  Conduit types: {len(conduit)}")
    print(f"  Wire types:    {len(wire)}")
    print(f"  Line items:    {len(line_items)}")
    print(f"  Total items:   {len(conduit) + len(wire) + len(line_items)}")
    print(f"\nOutput: {OUTPUT_PATH}")
    print("=" * 60)


if __name__ == "__main__":
    main()
