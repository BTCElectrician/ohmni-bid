#!/usr/bin/env python3
"""
Extract pricing data from the Excel master workbook into data/pricing_database.json.

Requirements:
- openpyxl
- msoffcrypto-tool (only if the workbook is password protected)
"""

from __future__ import annotations

import io
import json
import os
import sys
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[2]

try:
    from openpyxl import load_workbook
except ImportError as exc:
    print(f"Missing dependency: {exc}")
    print("Run: pip install openpyxl")
    sys.exit(1)

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None

if load_dotenv:
    load_dotenv(PROJECT_ROOT / ".env")

DEFAULT_XLSX = PROJECT_ROOT / "legacy" / "data" / "Unit price - Master 11-3-21- JG.xlsx"
EXCEL_PATH = Path(os.getenv("PRICING_XLSX_PATH", str(DEFAULT_XLSX)))
EXCEL_PASSWORD = os.getenv("EXCEL_PASSWORD", "")
OUTPUT_PATH = PROJECT_ROOT / "data" / "pricing_database.json"


def normalize_ref(value: str) -> str:
    text = str(value).upper().replace('"', "").replace("'", "")
    text = text.replace(" ", "_")
    return "".join(ch if ch.isalnum() or ch in "_-:" else "_" for ch in text)


def build_ref(*parts: str) -> str:
    return ":".join(normalize_ref(part) for part in parts if part)


def load_workbook_with_password(path: Path, password: str):
    if password and msoffcrypto:
        with open(path, "rb") as raw_file:
            office_file = msoffcrypto.OfficeFile(raw_file)
            office_file.load_key(password=password)
            decrypted = io.BytesIO()
            office_file.decrypt(decrypted)
            return load_workbook(decrypted, data_only=True)
    return load_workbook(path, data_only=True)


def safe_float(value):
    try:
        return float(value) if value is not None else 0.0
    except Exception:
        return 0.0


def extract_conduit_pricing(wb):
    conduit_items = []
    sheet_name = "FEEDERS"
    try:
        ws = wb[sheet_name]
    except KeyError:
        print("Warning: FEEDERS sheet not found")
        return conduit_items

    conduit_types = [
        {"type": "EMT_SS", "typeName": "EMT Set Screw", "start_row": 10, "end_row": 19},
        {"type": "EMT_COMP", "typeName": "EMT Compression", "start_row": 22, "end_row": 31},
        {"type": "HW", "typeName": "Heavy Wall/Rigid", "start_row": 34, "end_row": 43},
        {"type": "IMC", "typeName": "IMC", "start_row": 46, "end_row": 55},
        {"type": "PVC", "typeName": "PVC Schedule 40", "start_row": 58, "end_row": 68},
        {"type": "PVC_GRC", "typeName": "PVC/GRC Assembly", "start_row": 71, "end_row": 80},
    ]

    sizes = ["1/2", "3/4", "1", "1-1/4", "1-1/2", "2", "2-1/2", "3", "3-1/2", "4", "5"]

    for ct in conduit_types:
        row = ct["start_row"]
        size_idx = 0
        while row <= ct["end_row"] and size_idx < len(sizes):
            material_cost = ws.cell(row=row, column=4).value
            labor_hours = ws.cell(row=row, column=7).value

            if material_cost or labor_hours:
                conduit_items.append({
                    "externalRef": build_ref(sheet_name, "CONDUIT", ct["type"], f"R{row}"),
                    "type": ct["type"],
                    "typeName": ct["typeName"],
                    "name": f"{sizes[size_idx]}\" {ct['type'].replace('_', ' ')} Unit",
                    "size": sizes[size_idx],
                    "materialCostPer100ft": safe_float(material_cost),
                    "laborHoursPer100ft": safe_float(labor_hours),
                })

            row += 1
            size_idx += 1

    print(f"  Extracted {len(conduit_items)} conduit items")
    return conduit_items


def extract_wire_pricing(wb):
    wire_items = []
    sheet_name = "FEEDERS"
    try:
        ws = wb[sheet_name]
    except KeyError:
        return wire_items

    copper_sizes = [
        "#12", "#10", "#8", "#6", "#4", "#3", "#2", "#1",
        "#1/0", "#2/0", "#3/0", "#4/0",
        "250 MCM", "300 MCM", "350 MCM", "400 MCM", "500 MCM", "600 MCM", "750 MCM",
    ]

    for i, size in enumerate(copper_sizes):
        row = 83 + i
        if row > 101:
            break
        market_price = ws.cell(row=row, column=3).value
        labor_hours = ws.cell(row=row, column=7).value

        if market_price:
            if size in ["#12", "#10", "#8"]:
                markup_pct = 0.15
            elif size in ["#6", "#4"]:
                markup_pct = 0.10
            elif size == "750 MCM":
                markup_pct = 0.10
            else:
                markup_pct = 0.05

            wire_items.append({
                "externalRef": build_ref(sheet_name, "WIRE", "CU", size, f"R{row}"),
                "material": "CU",
                "type": "THHN",
                "size": size,
                "marketPricePer1000ft": safe_float(market_price),
                "markupPercent": markup_pct,
                "laborHoursPer1000ft": safe_float(labor_hours),
            })

    aluminum_sizes = ["#6", "#4", "#2", "#1/0", "#2/0", "#3/0", "#4/0",
                      "250 MCM", "300 MCM", "350 MCM", "400 MCM", "500 MCM", "600 MCM"]

    for i, size in enumerate(aluminum_sizes):
        row = 108 + i
        if row > 120:
            break
        market_price = ws.cell(row=row, column=3).value
        labor_hours = ws.cell(row=row, column=7).value

        if market_price:
            wire_items.append({
                "externalRef": build_ref(sheet_name, "WIRE", "AL", size, f"R{row}"),
                "material": "AL",
                "type": "THHN",
                "size": size,
                "marketPricePer1000ft": safe_float(market_price),
                "markupPercent": 0.05,
                "laborHoursPer1000ft": safe_float(labor_hours),
            })

    print(f"  Extracted {len(wire_items)} wire items")
    return wire_items


def extract_line_items(wb):
    line_items = []
    sheet_name = "basebid"
    try:
        ws = wb[sheet_name]
    except KeyError:
        print("Warning: basebid sheet not found")
        return line_items

    categories = [
        {"name": "TEMP_POWER", "start_row": 11, "end_row": 14},
        {"name": "ELECTRICAL_SERVICE", "start_row": 15, "end_row": 158},
        {"name": "MECHANICAL_CONNECTIONS", "start_row": 159, "end_row": 190},
        {"name": "INTERIOR_LIGHTING", "start_row": 191, "end_row": 231},
        {"name": "EXTERIOR_LIGHTING", "start_row": 232, "end_row": 247},
        {"name": "POWER_RECEPTACLES", "start_row": 248, "end_row": 285},
        {"name": "SITE_CONDUITS", "start_row": 286, "end_row": 294},
        {"name": "SECURITY", "start_row": 295, "end_row": 297},
        {"name": "FIRE_ALARM", "start_row": 299, "end_row": 311},
        {"name": "GENERAL_CONDITIONS", "start_row": 312, "end_row": 320},
    ]

    for cat in categories:
        for row in range(cat["start_row"], cat["end_row"] + 1):
            description = ws.cell(row=row, column=1).value
            material_cost = ws.cell(row=row, column=4).value
            unit_type = ws.cell(row=row, column=5).value
            labor_hours = ws.cell(row=row, column=7).value

            if not description or description == "MATERIAL":
                continue
            if not material_cost and not labor_hours:
                continue

            unit = str(unit_type).upper() if unit_type else "E"
            if unit == "LOT":
                unit = "Lot"
            elif unit not in ["E", "C", "M"]:
                unit = "E"

            line_items.append({
                "externalRef": build_ref(sheet_name, cat["name"], f"R{row}"),
                "category": cat["name"],
                "description": str(description).strip(),
                "materialUnitCost": safe_float(material_cost),
                "laborHoursPerUnit": safe_float(labor_hours),
                "unitType": unit,
            })

    print(f"  Extracted {len(line_items)} line items")
    return line_items


def extract_parameters(wb):
    params = {
        "laborRate": 118.0,
        "materialTaxRate": 0.1025,
        "defaultOverheadProfitRate": 0,
    }

    try:
        ws = wb["basebid"]
        labor_rate = ws.cell(row=323, column=9).value
        tax_rate = ws.cell(row=323, column=5).value
        op_rate = ws.cell(row=329, column=3).value

        if labor_rate:
            params["laborRate"] = safe_float(labor_rate)
        if tax_rate:
            params["materialTaxRate"] = safe_float(tax_rate)
        if op_rate:
            params["defaultOverheadProfitRate"] = safe_float(op_rate)
    except Exception as exc:
        print(f"Warning: Error extracting parameters: {exc}")

    print(f"  Parameters: ${params['laborRate']}/hr, {params['materialTaxRate'] * 100}% tax")
    return params


def main():
    print("=" * 60)
    print("OHMNI BID - Pricing Extraction (v2)")
    print("=" * 60)

    if not EXCEL_PATH.exists():
        print(f"\nError: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    print(f"\nSource: {EXCEL_PATH.name}")
    if EXCEL_PASSWORD:
        print("Password: (set)")
    else:
        print("Password: (none)")

    try:
        wb = load_workbook_with_password(EXCEL_PATH, EXCEL_PASSWORD)
    except Exception as exc:
        print(f"Error loading workbook: {exc}")
        sys.exit(1)

    print(f"Sheets found: {wb.sheetnames}")
    print("\nExtracting data...")

    parameters = extract_parameters(wb)
    conduit = extract_conduit_pricing(wb)
    wire = extract_wire_pricing(wb)
    line_items = extract_line_items(wb)

    output = {
        "version": "2.0",
        "source": EXCEL_PATH.name,
        "extracted_at": datetime.now().isoformat(),
        "parameters": parameters,
        "conduit": conduit,
        "wire": wire,
        "lineItems": line_items,
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w") as handle:
        json.dump(output, handle, indent=2)

    print("\n" + "=" * 60)
    print("EXTRACTION COMPLETE")
    print("=" * 60)
    print(f"  Conduit types: {len(conduit)}")
    print(f"  Wire types:    {len(wire)}")
    print(f"  Line items:    {len(line_items)}")
    print(f"  Total items:   {len(conduit) + len(wire) + len(line_items)}")
    print(f"\nOutput: {OUTPUT_PATH}")
    print("=" * 60)


if __name__ == "__main__":
    main()
